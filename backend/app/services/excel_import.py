import io
import uuid
from dataclasses import dataclass

import openpyxl
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Direction, Project, ProjectSource, TRLLevel
from app.schemas.excel_import import ImportPreviewResponse, ImportRowError

TEMPLATE_COLUMNS = [
    "Название проекта",
    "Предполагаемое проектное направление",
    "Проект уже реализуется в рамках проектной деятельности?",
    "Срок реализации проекта",
    "Актуальность",
    "Проблема",
    "Цель",
    "Ключевые задачи",
    "Ожидаемый продуктовый результат",
    "Уровень готовности технологий на текущий момент",
]

EXAMPLE_ROW = [
    "Пример проекта",
    "IT",
    "Нет",
    2,
    "Актуальность проекта",
    "Описание проблемы проекта",
    "Цель проекта",
    "Ключевые задачи проекта",
    "Ожидаемый продуктовый результат",
    3,
]


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"
    ws.append(TEMPLATE_COLUMNS)
    ws.append(EXAMPLE_ROW)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_bool(value: str | None) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in ("да", "yes", "1", "true")


def _parse_int(value: str | int | None, default: int = 0) -> int:
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


@dataclass
class ParsedRow:
    title: str
    direction_id: uuid.UUID | None
    is_ongoing: bool
    implementation_period: int
    relevance: str | None
    problem: str | None
    goal: str | None
    key_tasks: str | None
    expected_result: str | None
    trl_id: uuid.UUID | None


class ExcelImportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self._directions: dict[str, uuid.UUID] = {}
        self._trl_by_level: dict[int, uuid.UUID] = {}

    async def _load_dicts(self) -> None:
        directions = (
            await self.session.execute(select(Direction).where(Direction.is_active.is_(True)))
        ).scalars().all()
        self._directions = {d.name.strip().lower(): d.id for d in directions}

        trl_levels = (
            await self.session.execute(select(TRLLevel).where(TRLLevel.is_active.is_(True)))
        ).scalars().all()
        self._trl_by_level = {t.level: t.id for t in trl_levels}

    async def _get_existing_titles(self) -> set[str]:
        result = await self.session.execute(select(Project.title))
        return {row[0].strip().lower() for row in result.fetchall()}

    def _parse_file(self, file_bytes: bytes) -> list[dict]:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []

        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        col_idx = {name: i for i, name in enumerate(header)}

        def get_cell(row: tuple, col_name: str) -> str | None:
            idx = col_idx.get(col_name)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            if val is None:
                return None
            text = str(val).strip()
            return text if text else None

        def get_raw(row: tuple, col_name: str):
            idx = col_idx.get(col_name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        data = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            data.append(
                {
                    "title": get_cell(row, "Название проекта"),
                    "direction_name": get_cell(row, "Предполагаемое проектное направление"),
                    "is_ongoing_raw": get_cell(
                        row, "Проект уже реализуется в рамках проектной деятельности?"
                    ),
                    "implementation_period_raw": get_raw(row, "Срок реализации проекта"),
                    "relevance": get_cell(row, "Актуальность"),
                    "problem": get_cell(row, "Проблема"),
                    "goal": get_cell(row, "Цель"),
                    "key_tasks": get_cell(row, "Ключевые задачи"),
                    "expected_result": get_cell(row, "Ожидаемый продуктовый результат"),
                    "trl_level_raw": get_raw(
                        row, "Уровень готовности технологий на текущий момент"
                    ),
                }
            )
        return data

    async def parse_and_validate(
        self, file_bytes: bytes
    ) -> tuple[list[ParsedRow], ImportPreviewResponse]:
        await self._load_dicts()
        existing_titles = await self._get_existing_titles()

        raw_rows = self._parse_file(file_bytes)

        errors: list[ImportRowError] = []
        duplicates: list[str] = []
        valid_rows: list[ParsedRow] = []
        seen_titles: set[str] = set()
        seen_duplicates: set[str] = set()

        for i, raw in enumerate(raw_rows, start=2):  # row 1 is header
            row_errors: list[ImportRowError] = []

            title = raw.get("title")
            if not title:
                errors.append(
                    ImportRowError(row=i, field="Название проекта", message="Обязательное поле")
                )
                continue

            title_lower = title.strip().lower()

            if title_lower in seen_titles or title_lower in existing_titles:
                if title_lower not in seen_duplicates:
                    seen_duplicates.add(title_lower)
                    duplicates.append(title)
            seen_titles.add(title_lower)

            direction_id = None
            if raw.get("direction_name"):
                direction_id = self._directions.get(raw["direction_name"].lower())
                if direction_id is None:
                    row_errors.append(
                        ImportRowError(
                            row=i,
                            field="Направление",
                            message=f"Значение '{raw['direction_name']}' не найдено в справочнике",
                        )
                    )

            trl_id = None
            trl_raw = raw.get("trl_level_raw")
            if trl_raw is not None:
                trl_num = _parse_int(trl_raw)
                trl_id = self._trl_by_level.get(trl_num)
                if trl_id is None:
                    row_errors.append(
                        ImportRowError(
                            row=i,
                            field="УГТ",
                            message=f"Уровень {trl_raw} не найден в справочнике",
                        )
                    )

            if row_errors:
                errors.extend(row_errors)
                continue

            valid_rows.append(
                ParsedRow(
                    title=title,
                    direction_id=direction_id,
                    is_ongoing=_parse_bool(raw.get("is_ongoing_raw")),
                    implementation_period=_parse_int(
                        raw.get("implementation_period_raw"), default=0
                    ),
                    relevance=raw.get("relevance"),
                    problem=raw.get("problem"),
                    goal=raw.get("goal"),
                    key_tasks=raw.get("key_tasks"),
                    expected_result=raw.get("expected_result"),
                    trl_id=trl_id,
                )
            )

        preview = ImportPreviewResponse(
            valid_count=len(valid_rows),
            error_count=len(errors),
            errors=errors,
            duplicates=duplicates,
        )
        return valid_rows, preview

    async def do_import(self, valid_rows: list[ParsedRow]) -> int:
        projects = [
            Project(
                title=row.title[:80],
                direction_id=row.direction_id,
                is_ongoing=row.is_ongoing,
                implementation_period=row.implementation_period,
                relevance=row.relevance or "",
                problem=row.problem or "",
                goal=row.goal or "",
                key_tasks=row.key_tasks or "",
                expected_result=row.expected_result or "",
                trl_id=row.trl_id,
                source=ProjectSource.auto,
            )
            for row in valid_rows
        ]
        self.session.add_all(projects)
        await self.session.flush()  # populate project.id before commit
        project_ids = [str(p.id) for p in projects]
        await self.session.commit()

        # Trigger async embedding generation for newly imported projects
        from app.tasks.embeddings import bulk_generate_embeddings

        bulk_generate_embeddings.delay(project_ids)

        return len(projects)
