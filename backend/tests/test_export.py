import io
import uuid
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest

from app.database import get_db
from app.main import app
from app.models import ProjectSource
from app.services.excel_export import ExcelExportService


def _make_project(
    title: str = "Тест",
    group_id: uuid.UUID | None = None,
    group_name: str | None = None,
    is_selected: bool = False,
    is_ongoing: bool = False,
    source: ProjectSource = ProjectSource.manual,
) -> MagicMock:
    p = MagicMock()
    p.title = title
    p.problem = "Проблема"
    p.goal = "Цель"
    p.expected_result = "Результат"
    p.is_ongoing = is_ongoing
    p.is_selected = is_selected
    p.source = source
    p.direction = None
    p.priority_direction = None
    p.trl_level = None
    p.group_id = group_id
    if group_name:
        p.group = MagicMock()
        p.group.name = group_name
    else:
        p.group = None
    return p


def _read_xlsx(content: bytes) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


class TestExcelExportService:
    def test_build_xlsx_returns_valid_xlsx(self):
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("Проект 1"), _make_project("Проект 2")]
        result = service._build_xlsx(projects, {})
        assert result[:4] == b"PK\x03\x04"  # xlsx magic bytes

    def test_build_xlsx_columns(self):
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("Проект А")]
        result = service._build_xlsx(projects, {})
        rows = _read_xlsx(result)
        header = rows[0]
        assert "Название" in header
        assert "Группа" in header
        assert "Score" in header

    def test_build_xlsx_group_info(self):
        group_id = uuid.uuid4()
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("Проект", group_id=group_id, group_name="Группа 1")]
        avg_scores = {group_id: 0.85}
        result = service._build_xlsx(projects, avg_scores)
        rows = _read_xlsx(result)
        header = rows[0]
        data_row = rows[1]
        group_col = header.index("Группа")
        score_col = header.index("Score")
        assert data_row[group_col] == "Группа 1"
        assert data_row[score_col] == pytest.approx(0.85)

    def test_build_xlsx_no_group(self):
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("Без группы")]
        result = service._build_xlsx(projects, {})
        rows = _read_xlsx(result)
        header = rows[0]
        data_row = rows[1]
        group_col = header.index("Группа")
        score_col = header.index("Score")
        assert data_row[group_col] is None
        assert data_row[score_col] is None

    def test_build_xlsx_is_ongoing(self):
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("Текущий", is_ongoing=True)]
        result = service._build_xlsx(projects, {})
        rows = _read_xlsx(result)
        header = rows[0]
        data_row = rows[1]
        col = header.index("Текущий")
        assert data_row[col] == "Да"

    def test_build_xlsx_empty(self):
        service = ExcelExportService(AsyncMock())
        result = service._build_xlsx([], {})
        rows = _read_xlsx(result)
        # header only
        assert len(rows) == 1

    @pytest.mark.asyncio
    async def test_export_all_context(self):
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("П1"), _make_project("П2")]
        service._fetch_all_projects = AsyncMock(return_value=projects)
        service._get_group_avg_scores = AsyncMock(return_value={})
        result = await service.export("all")
        assert result[:4] == b"PK\x03\x04"
        service._fetch_all_projects.assert_called_once()

    @pytest.mark.asyncio
    async def test_export_selected_context(self):
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("Отобранный", is_selected=True)]
        service._fetch_selected_projects = AsyncMock(return_value=projects)
        service._get_group_avg_scores = AsyncMock(return_value={})
        result = await service.export("selected")
        assert result[:4] == b"PK\x03\x04"
        service._fetch_selected_projects.assert_called_once()

    @pytest.mark.asyncio
    async def test_export_grouped_context(self):
        group_id = uuid.uuid4()
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("Сгруппированный", group_id=group_id, group_name="Г1")]
        service._fetch_grouped_projects = AsyncMock(return_value=projects)
        service._get_group_avg_scores = AsyncMock(return_value={group_id: 0.9})
        result = await service.export("grouped")
        service._fetch_grouped_projects.assert_called_once()
        rows = _read_xlsx(result)
        header = rows[0]
        score_col = header.index("Score")
        assert rows[1][score_col] == pytest.approx(0.9)

    @pytest.mark.asyncio
    async def test_export_filtered_context_uses_provided_projects(self):
        service = ExcelExportService(AsyncMock())
        projects = [_make_project("Отфильтрованный")]
        service._get_group_avg_scores = AsyncMock(return_value={})
        result = await service.export("filtered", filtered_projects=projects)
        rows = _read_xlsx(result)
        assert rows[1][0] == "Отфильтрованный"


@pytest.mark.asyncio
async def test_export_endpoint_returns_xlsx():
    mock_db = AsyncMock()

    async def mock_export(context, filtered_projects=None):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Название"])
        buf = __import__("io").BytesIO()
        wb.save(buf)
        return buf.getvalue()

    with patch("app.api.projects.ExcelExportService") as MockSvc:
        MockSvc.return_value.export = mock_export
        app.dependency_overrides[get_db] = lambda: mock_db
        try:
            from httpx import ASGITransport, AsyncClient

            async with AsyncClient(
                transport=ASGITransport(app=app), base_url="http://test"
            ) as client:
                response = await client.get("/api/projects/export")
            assert response.status_code == 200
            assert "attachment" in response.headers["content-disposition"]
            assert "projects_export.xlsx" in response.headers["content-disposition"]
            assert response.content[:4] == b"PK\x03\x04"
        finally:
            app.dependency_overrides.pop(get_db, None)


@pytest.mark.asyncio
async def test_export_endpoint_selected_context():
    mock_db = AsyncMock()

    async def mock_export(context, filtered_projects=None):
        assert context == "selected"
        from openpyxl import Workbook

        wb = Workbook()
        buf = __import__("io").BytesIO()
        wb.save(buf)
        return buf.getvalue()

    with patch("app.api.projects.ExcelExportService") as MockSvc:
        MockSvc.return_value.export = mock_export
        app.dependency_overrides[get_db] = lambda: mock_db
        try:
            from httpx import ASGITransport, AsyncClient

            async with AsyncClient(
                transport=ASGITransport(app=app), base_url="http://test"
            ) as client:
                response = await client.get("/api/projects/export?context=selected")
            assert response.status_code == 200
        finally:
            app.dependency_overrides.pop(get_db, None)


@pytest.mark.asyncio
async def test_export_endpoint_grouped_context():
    mock_db = AsyncMock()

    async def mock_export(context, filtered_projects=None):
        assert context == "grouped"
        from openpyxl import Workbook

        wb = Workbook()
        buf = __import__("io").BytesIO()
        wb.save(buf)
        return buf.getvalue()

    with patch("app.api.projects.ExcelExportService") as MockSvc:
        MockSvc.return_value.export = mock_export
        app.dependency_overrides[get_db] = lambda: mock_db
        try:
            from httpx import ASGITransport, AsyncClient

            async with AsyncClient(
                transport=ASGITransport(app=app), base_url="http://test"
            ) as client:
                response = await client.get("/api/projects/export?context=grouped")
            assert response.status_code == 200
        finally:
            app.dependency_overrides.pop(get_db, None)
